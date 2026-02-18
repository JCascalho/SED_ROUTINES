"""
====================================================================================
Tracer metric processing script
File: tracer_metrics_processing.py
====================================================================================

Author: João Cascalho (FCUL / IDL – University of Lisbon)
Date: December 2025

Purpose
-------
This script processes seabed fluorescent tracer monitoring data and computes
a suite of tracer-transport and recovery metrics, including:

    • Calibrated tracer mass concentration (TMcal)
    • Tracer centre-of-mass (TCOM) displacement
    • Horizontal diffusion coefficients following Miller & Komar (1979)
    • Mean tracer transport velocities
    • Horizontal tracer transport fluxes
    • Number of tracer grains per square metre (per sample)
    • Tracer mass recovery within Thiessen polygons generated in ArcGIS Pro
      (T_rate, kg)
    • Percentage of injected tracer mass recovered (T_rate_%)
    • A summary table of tracer recovery across campaigns (including cumulative %)
    • Summary diffusion and transport tables (scaled ×10⁻⁸, SI units)

Scientific Background
---------------------
Field monitoring images of fluorescent tracer grains are used to derive the
area concentration (Ac). Ac is defined as the ratio between the number of pixels
identified as tracer and the total number of pixels within a fixed photo area:

    Ac = (number of tracer-particle pixels) / (total pixel count of photo area)

Because camera geometry and image footprint remain constant across sampling
events, Ac provides a spatially normalized, dimensionless proxy for tracer
abundance.

Laboratory calibration experiments establish a linear relationship between Ac
and the true tracer mass concentration (TMc):

    Ac = m · TMc        (R² ≈ 0.99 in the calibration example)

where:
    m = slope of the calibration regression line.

Rearranging yields the calibrated tracer mass concentration used in this study:

    TMcal = TMc = Ac / m

In addition, the script computes:
    • Tracer grain density per unit area:
          N_grains/m² = tracer_nr / photo_area
    • Tracer mass recovery per Thiessen polygon:
          T_rate = TMcal · Tp_area · dreger_d · rho_sed · sed_poro

      where:
          TMcal     = calibrated tracer mass concentration
          Tp_area   = Thiessen polygon area [m²]
          dreger_d  = mean drag penetration depth [m]
          rho_sed   = sediment density [kg/m³]
          sed_poro  = sediment porosity [-]

The total injected tracer mass is tracer_m (≈ 600 kg). The percentage of tracer
mass recovered per polygon is computed as:

    T_rate_% = 100 · T_rate / tracer_m

Input Data Assumptions
----------------------
The input Excel workbook is assumed to contain:
    • A reference sheet (CM0) with coordinates, dates, and global constants
    • Campaign sheets (C2–C7) with tracer observations, geometry, and calibration
      slope (m)
    • Columns Start_date and End_date formatted as datetime objects
    • Consistent coordinate systems and units across all sheets
    • Non-zero tracer mass concentration for centre-of-mass and diffusion
      calculations

Output
------
The script generates:

    • Cleaned and augmented campaign sheets (C2–C7) with derived tracer metrics
      (including grains/m², T_rate, and T_rate_%)
    • CM2–CM7 summary sheets with mean grains/m² and summed T_rate and T_rate_%
    • A diffusion coefficient table (scaled ×10⁻⁸, SI units, 2 decimal places)
    • Mean velocity and transport rate tables (scaled ×10⁻⁸, SI units)
    • A tracer recovery summary sheet with cumulative percentage recovery
    • A PNG figure ("Tracer_Recovery_Histogram.png") showing tracer recovery
      percentages for each sampling campaign

====================================================================================
"""

import pandas as pd
import math
from openpyxl import load_workbook


def scaled_1e8(val: float) -> float:
    """
    Scale value by 1e8 and round to 2 decimals.
    Keep NaN as NaN (Excel will show empty).
    """
    if pd.isna(val):
        return float("nan")
    return round(val * 1e8, 2)


def main():
    # -------------------------------------------------
    # 0. File names (change if needed)
    # -------------------------------------------------
    input_file = "input_TRACER_DATA.xlsx"
    output_file = "output_TRACER_DATA_METRICS.xlsx"

    # -------------------------------------------------
    # 1. Load workbook and read CM0 (reference sheet)
    # -------------------------------------------------
    xls = pd.ExcelFile(input_file)

    if "CM0" not in xls.sheet_names:
        raise ValueError("Sheet 'CM0' not found! It is required as reference.")

    cM0 = pd.read_excel(
        xls, sheet_name="CM0", parse_dates=["Start_date", "End_date"]
    )

    # Reference end date and CM0 coordinates (for first Δt and first displacement)
    previous_end_date = cM0.loc[0, "End_date"]
    prev_CMx = cM0.loc[0, "x"]
    prev_CMy = cM0.loc[0, "y"]
    prev_CMas = cM0.loc[0, "as"]
    prev_CMcs = cM0.loc[0, "cs"]

    # ---- global constants (read from CM0 if present, else defaults) ----
    # drag penetration depth [m]
    if "dreger_d" in cM0.columns:
        dreger_d_global = float(cM0["dreger_d"].iloc[0])
    else:
        dreger_d_global = 0.15

    # sediment porosity [-]
    if "sed_poro" in cM0.columns:
        sed_poro_global = float(cM0["sed_poro"].iloc[0])
    else:
        sed_poro_global = 0.6

    # sediment density [kg/m3] – will be overridden per sheet if present
    if "rho_sed" in cM0.columns:
        rho_sed_global = float(cM0["rho_sed"].iloc[0])
    else:
        rho_sed_global = float("nan")

    # injected tracer mass [kg]
    if "tracer_m" in cM0.columns:
        tracer_m_total = float(cM0["tracer_m"].iloc[0])
    else:
        tracer_m_total = 600.0

    # -------------------------------------------------
    # 2. Define the sheets to process (C2 to C7)
    # -------------------------------------------------
    target_sheets = [f"C{i}" for i in range(2, 8)]  # C2, C3, C4, C5, C6, C7

    # Containers to store results
    processed_sheets = {}   # updated C2..C7
    com_sheets = {}         # CM2..CM7
    vmq_rows = []           # rows for Vm/Q summary table (scaled numeric)
    diff_rows = []          # rows for diffusion summary table (scaled numeric)
    prev_label = "C0"       # for interval labels C0–C2, C2–C3, ...

    # -------------------------------------------------
    # 3. Loop over C2–C7 and compute metrics
    # -------------------------------------------------
    for sheet_name in target_sheets:

        if sheet_name not in xls.sheet_names:
            print(f"⚠ WARNING: Sheet '{sheet_name}' not found in workbook. Skipping.")
            continue

        print(f"\n📌 Processing sheet: {sheet_name}")

        df = pd.read_excel(
            xls, sheet_name=sheet_name, parse_dates=["Start_date", "End_date"]
        )

        if "End_date" not in df.columns:
            raise KeyError(f"'End_date' column not found in sheet '{sheet_name}'.")

        # End_date for this campaign (assume representative = first row)
        this_end_date = df["End_date"].iloc[0]

        # ---------------------------------------------
        # 3.1 Compute Δt [seconds] between SUCCESSIVE CM's
        # ---------------------------------------------
        delta_t_seconds = (this_end_date - previous_end_date).total_seconds()
        df["delta_t"] = delta_t_seconds

        # ---------------------------------------------
        # 3.1a Tracer grains per square metre
        #       grains_per_m2 = tracer_nr / photo_area
        # ---------------------------------------------
        if "tracer_nr" in df.columns and "photo_area" in df.columns:
            df["tracer_grains_m2"] = df["tracer_nr"] / df["photo_area"]
        else:
            print(
                f"⚠ WARNING: 'tracer_nr' or 'photo_area' not found in '{sheet_name}'. "
                "Column 'tracer_grains_m2' will not be created."
            )
            df["tracer_grains_m2"] = float("nan")

        # ---------------------------------------------
        # 3.2 Compute TMcal = Ac / m
        # ---------------------------------------------
        if "Ac " in df.columns:
            ac_col = "Ac "
        elif "Ac" in df.columns:
            ac_col = "Ac"
        else:
            raise KeyError(f"Missing 'Ac' or 'Ac ' column in sheet '{sheet_name}'.")

        df["TMcal"] = df[ac_col] / df["m"]

        # ---------------------------------------------
        # 3.2a Tracer mass recuperation per polygon (T_rate, kg)
        #       T_rate = TMcal * Tp_area * dreger_d * rho_sed * sed_poro
        # ---------------------------------------------
        # Thiessen polygon area
        if "Tp_area" in df.columns:
            tp_area_col = "Tp_area"
        else:
            raise KeyError(
                f"Missing 'Tp_area' column in sheet '{sheet_name}' "
                "(Thiessen polygon area required for T_rate)."
            )

        # get campaign-specific constants if present, else global
        if "dreger_d" in df.columns:
            dreger_d = float(df["dreger_d"].iloc[0])
        else:
            dreger_d = dreger_d_global

        if "sed_poro" in df.columns:
            sed_poro = float(df["sed_poro"].iloc[0])
        else:
            sed_poro = sed_poro_global

        if "rho_sed" in df.columns:
            rho_sed = float(df["rho_sed"].iloc[0])
        else:
            if math.isnan(rho_sed_global):
                raise KeyError(
                    f"'rho_sed' (sediment density) not found in sheet '{sheet_name}' "
                    "or in CM0. Please add it to the input file."
                )
            rho_sed = rho_sed_global

        # tracer mass injected (same for all campaigns)
        tracer_m = tracer_m_total

        df["T_rate"] = df["TMcal"] * df[tp_area_col] * dreger_d * rho_sed * sed_poro
        df["T_rate_%"] = 100.0 * df["T_rate"] / tracer_m

        # ---------------------------------------------
        # 3.3 Distances from previous center of mass (per sample)
        #      Dist_x(i) = x_i - CMx_prev, etc.
        # ---------------------------------------------
        df["Dist_x"] = df["x"] - prev_CMx
        df["Dist_y"] = df["y"] - prev_CMy
        df["Dist_as"] = df["as"] - prev_CMas
        df["Dist_cs"] = df["cs"] - prev_CMcs

        # For checking like your Excel example:
        df["TMcal*Dist_x"] = df["TMcal"] * df["Dist_x"]
        df["TMcal*Dist_y"] = df["TMcal"] * df["Dist_y"]
        df["TMcal*Dist_as"] = df["TMcal"] * df["Dist_as"]
        df["TMcal*Dist_cs"] = df["TMcal"] * df["Dist_cs"]

        df["(TMcal*Dist_x)^2"] = df["TMcal*Dist_x"] ** 2
        df["(TMcal*Dist_y)^2"] = df["TMcal*Dist_y"] ** 2
        df["(TMcal*Dist_as)^2"] = df["TMcal*Dist_as"] ** 2
        df["(TMcal*Dist_cs)^2"] = df["TMcal*Dist_cs"] ** 2

        # ---------------------------------------------
        # 3.4 TMcal-weighted coordinates for COM
        # ---------------------------------------------
        df["TMcal*x"] = df["TMcal"] * df["x"]
        df["TMcal*y"] = df["TMcal"] * df["y"]
        df["TMcal*as"] = df["TMcal"] * df["as"]
        df["TMcal*cs"] = df["TMcal"] * df["cs"]

        # ---------------------------------------------
        # 3.5 Compute Center of Mass for this sheet
        # ---------------------------------------------
        TM_sum = df["TMcal"].sum()
        TM_sq_sum = (df["TMcal"] ** 2).sum()

        # Initialise all metrics
        D_x = D_y = D_as = D_cs = float("nan")
        Dr_xy = Dr_as_cs = float("nan")
        Vm_x = Vm_y = Vm_as = Vm_cs = float("nan")
        Q_x = Q_y = Q_as = Q_cs = float("nan")
        delta_mix_val = float("nan")
        grains_m2_mean = float("nan")
        T_rate_sum = float("nan")
        T_rate_pct_sum = float("nan")

        if TM_sum == 0 or math.isclose(TM_sum, 0.0):
            print(
                f"⚠ WARNING: Sum(TMcal) = 0 in '{sheet_name}'. "
                "COM, diffusion & transport set to NaN."
            )
            CMx = CMy = CMas = CMcs = float("nan")
        else:
            # ----- Center of mass -----
            CMx = df["TMcal*x"].sum() / TM_sum
            CMy = df["TMcal*y"].sum() / TM_sum
            CMas = df["TMcal*as"].sum() / TM_sum
            CMcs = df["TMcal*cs"].sum() / TM_sum

            # Mean grains/m2 in this campaign (simple arithmetic mean)
            grains_m2_mean = df["tracer_grains_m2"].mean()

            # Total tracer mass recovered in this campaign
            T_rate_sum = df["T_rate"].sum()
            T_rate_pct_sum = 100.0 * T_rate_sum / tracer_m

            # -----------------------------------------
            # 3.6 Diffusion coefficients (Miller & Komar, per-sample Dist_x)
            #     D_x = Σ(TMcal * Dist_x)^2 / (Σ TMcal^2 * Δt)
            # -----------------------------------------
            if delta_t_seconds == 0 or TM_sq_sum == 0 or math.isclose(TM_sq_sum, 0.0):
                print(
                    f"⚠ WARNING: Zero denominator / Δt in '{sheet_name}'. "
                    "Diffusion set to NaN."
                )
            else:
                num_x = df["(TMcal*Dist_x)^2"].sum()
                num_y = df["(TMcal*Dist_y)^2"].sum()
                num_as = df["(TMcal*Dist_as)^2"].sum()
                num_cs = df["(TMcal*Dist_cs)^2"].sum()

                denom = TM_sq_sum * delta_t_seconds

                D_x = num_x / denom
                D_y = num_y / denom
                D_as = num_as / denom
                D_cs = num_cs / denom

                Dr_xy = math.sqrt(D_x**2 + D_y**2)
                Dr_as_cs = math.sqrt(D_as**2 + D_cs**2)

            # -----------------------------------------
            # 3.7 Mean tracer displacement velocities
            #     based on COM displacement
            # -----------------------------------------
            dist_cm_x = CMx - prev_CMx
            dist_cm_y = CMy - prev_CMy
            dist_cm_as = CMas - prev_CMas
            dist_cm_cs = CMcs - prev_CMcs

            if delta_t_seconds == 0:
                print(
                    f"⚠ WARNING: Δt = 0 in '{sheet_name}'. "
                    "Velocities set to NaN."
                )
            else:
                Vm_x = dist_cm_x / delta_t_seconds
                Vm_y = dist_cm_y / delta_t_seconds
                Vm_as = dist_cm_as / delta_t_seconds
                Vm_cs = dist_cm_cs / delta_t_seconds

            # -----------------------------------------
            # 3.8 Transport rates Qx, Qy, Qas, Qcs
            # -----------------------------------------
            if "delta_mix" in df.columns:
                delta_mix_col = "delta_mix"
            elif "δ_mix" in df.columns:
                delta_mix_col = "δ_mix"
            elif "delta_mix " in df.columns:
                delta_mix_col = "delta_mix "
            else:
                delta_mix_col = None

            if delta_mix_col is None:
                print(
                    f"⚠ WARNING: No 'delta_mix' column found in '{sheet_name}'. "
                    "Transport Q set to NaN."
                )
            else:
                delta_mix_series = df[delta_mix_col].dropna()
                if delta_mix_series.empty:
                    print(
                        f"⚠ WARNING: 'delta_mix' column empty in '{sheet_name}'. "
                        "Transport Q set to NaN."
                    )
                else:
                    delta_mix_val = float(delta_mix_series.iloc[0])
                    Q_x = Vm_x * delta_mix_val
                    Q_y = Vm_y * delta_mix_val
                    Q_as = Vm_as * delta_mix_val
                    Q_cs = Vm_cs * delta_mix_val

        # Store updated C* sheet
        processed_sheets[sheet_name] = df

        # ---------------------------------------------
        # 3.9 Build corresponding CM* sheet (CM2, CM3, ...)
        # ---------------------------------------------
        cm_sheet_name = "CM" + sheet_name[1:]     # e.g. C2 -> CM2
        cm_value = cm_sheet_name                  # value in first column (CM2, CM3, ...)

        # COM displacement (for reference)
        dist_cm_x = CMx - prev_CMx
        dist_cm_y = CMy - prev_CMy
        dist_cm_as = CMas - prev_CMas
        dist_cm_cs = CMcs - prev_CMcs

        cm_df = pd.DataFrame({
            "CM":                    [cm_value],          # CM2, CM3, ..., CM7
            "CM_x":                  [CMx],
            "CM_y":                  [CMy],
            "CM_as":                 [CMas],
            "CM_cs":                 [CMcs],
            "TMcal_sum":             [TM_sum],
            "TMcal_sq_sum":          [TM_sq_sum],
            "delta_t_s":             [delta_t_seconds],   # time between this CM and previous CM
            "DistCM_x":              [dist_cm_x],
            "DistCM_y":              [dist_cm_y],
            "DistCM_as":             [dist_cm_as],
            "DistCM_cs":             [dist_cm_cs],

            "TCOMd_xy":              [math.sqrt(dist_cm_x**2 + dist_cm_y**2)],
            "TCOMd_ascs":            [math.sqrt(dist_cm_as**2 + dist_cm_cs**2)],
            "TCOMdir_deg_xy":        [(math.degrees(math.atan2(dist_cm_x, dist_cm_y)) + 360) % 360],
            "TCOMdir_deg_ascs":      [(math.degrees(math.atan2(dist_cm_as, dist_cm_cs)) + 360) % 360],
            "D_x":                   [D_x],
            "D_y":                   [D_y],
            "D_as":                  [D_as],
            "D_cs":                  [D_cs],
            "Dr_xy":                 [Dr_xy],
            "Dr_as_cs":              [Dr_as_cs],
            "Vm_x":                  [Vm_x],
            "Vm_y":                  [Vm_y],
            "Vm_as":                 [Vm_as],
            "Vm_cs":                 [Vm_cs],
            "delta_mix":             [delta_mix_val],
            "Q_x":                   [Q_x],
            "Q_y":                   [Q_y],
            "Q_as":                  [Q_as],
            "Q_cs":                  [Q_cs],
            "tracer_grains_m2_mean": [grains_m2_mean],
            "T_rate_sum":            [T_rate_sum],       # total kg in this campaign
            "T_rate_%_sum":          [T_rate_pct_sum],   # % of injected mass
            "prev_EndDate":          [previous_end_date],
            "End_date":              [this_end_date],
        })

        com_sheets[cm_sheet_name] = cm_df

        # ---------------------------------------------
        # 3.10 Collect Vm/Q row for summary table (scaled ×1e8, numeric)
        #      NOW including resultants of Qxy and Qas_cs
        # ---------------------------------------------
        interval_label = f"{prev_label}–{sheet_name}"  # e.g. C0–C2, C2–C3, ...

        # Compute unscaled resultants, guarding against NaNs
        if any(math.isnan(v) for v in (Q_x, Q_y)):
            Q_xy_res = float("nan")
        else:
            Q_xy_res = math.sqrt(Q_x**2 + Q_y**2)

        if any(math.isnan(v) for v in (Q_as, Q_cs)):
            Q_as_cs_res = float("nan")
        else:
            Q_as_cs_res = math.sqrt(Q_as**2 + Q_cs**2)

        vmq_rows.append({
            "Sampling intervals":    interval_label,
            "Vmx":                   scaled_1e8(Vm_x),
            "Vmy":                   scaled_1e8(Vm_y),
            "Vmas":                  scaled_1e8(Vm_as),
            "Vmcs":                  scaled_1e8(Vm_cs),
            "Qx":                    scaled_1e8(Q_x),
            "Qy":                    scaled_1e8(Q_y),
            "Qas":                   scaled_1e8(Q_as),
            "Qcs":                   scaled_1e8(Q_cs),
            "Q_xy_resultant":        scaled_1e8(Q_xy_res),
            "Q_as_cs_resultant":     scaled_1e8(Q_as_cs_res),
        })

        # ---------------------------------------------
        # 3.11 Collect diffusion row for summary table (scaled ×1e8, numeric)
        # ---------------------------------------------
        diff_rows.append({
            "Sampling intervals": interval_label,
            "Dx":       scaled_1e8(D_x),
            "Dy":       scaled_1e8(D_y),
            "Das":      scaled_1e8(D_as),
            "Dcs":      scaled_1e8(D_cs),
            "Dr_xy":    scaled_1e8(Dr_xy),
            "Dr_as_cs": scaled_1e8(Dr_as_cs),
        })

        # Update previous_end_date, previous CM coordinates and label for the *next* sheet
        previous_end_date = this_end_date
        prev_CMx = CMx
        prev_CMy = CMy
        prev_CMas = CMas
        prev_CMcs = CMcs
        prev_label = sheet_name

    # -------------------------------------------------
    # 4. Build Vm/Q summary DataFrame (scaled numeric)
    # -------------------------------------------------
    vmq_df = pd.DataFrame(
        vmq_rows,
        columns=[
            "Sampling intervals",
            "Vmx", "Vmy", "Vmas", "Vmcs",
            "Qx", "Qy", "Qas", "Qcs",
            "Q_xy_resultant", "Q_as_cs_resultant",
        ]
    )

    # -------------------------------------------------
    # 5. Build Diffusion summary DataFrame (scaled numeric)
    # -------------------------------------------------
    diff_df = pd.DataFrame(
        diff_rows,
        columns=[
            "Sampling intervals",
            "Dx", "Dy", "Das", "Dcs",
            "Dr_xy", "Dr_as_cs",
        ]
    )

    # -------------------------------------------------
    # 6. Build Tracer Mass Recovery Summary Table
    # -------------------------------------------------
    recovery_rows = []

    # Ensure campaigns are in chronological order CM2, CM3, ... CM7
    campaign_order = sorted(
        com_sheets.keys(),
        key=lambda s: int(s[2:])  # CM2 -> 2, CM3 -> 3, etc.
    )

    for cm_name in campaign_order:
        cm_df = com_sheets[cm_name]
        row = cm_df.iloc[0]
        recovery_rows.append({
            "Campaign": cm_name,
            "T_rate_sum_kg": float(row["T_rate_sum"]),
            "T_rate_%": float(row["T_rate_%_sum"]),
        })

    # DataFrame with one row per CM campaign
    recovery_df = pd.DataFrame(recovery_rows)

    # Cumulative % recovery (per campaign)
    recovery_df["T_rate_%_cum"] = recovery_df["T_rate_%"].cumsum()

    # Add final TOTAL row
    total_T = recovery_df["T_rate_sum_kg"].sum()
    total_pct = recovery_df["T_rate_%"].sum()
    total_cum = recovery_df["T_rate_%_cum"].iloc[-1]

    recovery_df.loc[len(recovery_df.index)] = {
        "Campaign": "TOTAL",
        "T_rate_sum_kg": total_T,
        "T_rate_%": total_pct,
        "T_rate_%_cum": total_cum,
    }

    # -------------------------------------------------
    # 7. Write everything to the output workbook
    # -------------------------------------------------
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        # Write CM0 unchanged
        cM0.to_excel(writer, sheet_name="CM0", index=False)

        # Write processed C2..C7 sheets
        for name, df in processed_sheets.items():
            df.to_excel(writer, sheet_name=name, index=False)

        # Write CM2..CM7 sheets (one row each)
        for cm_name, cm_df in com_sheets.items():
            cm_df.to_excel(writer, sheet_name=cm_name, index=False)

        # Write Vm/Q summary table (scaled ×10^-8, numeric)
        vmq_df.to_excel(writer, sheet_name="Vm_Q_table_10^(-8)", index=False)

        # Write diffusion summary table (scaled ×10^-8, numeric)
        diff_df.to_excel(writer, sheet_name="Diffusion_table_10^(-8)", index=False)

        # Write tracer mass recovery summary
        recovery_df.to_excel(
            writer,
            sheet_name="Tracer_Recovery_Summary",
            index=False
        )

    # -------------------------------------------------
    # 8. Apply Excel number format "0.00" to summary tables
    #     (keeps values numeric but always shows 2 decimals)
    # -------------------------------------------------
    wb = load_workbook(output_file)

    for sheet_name in [
        "Vm_Q_table_10^(-8)",
        "Diffusion_table_10^(-8)",
        "Tracer_Recovery_Summary",
    ]:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Start from row 2 (skip header), col 2 (skip first text column)
            for row in ws.iter_rows(min_row=2, min_col=2):
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "0.00"

    # -------------------------------------------------
    # 9. Generate tracer recovery histogram by campaign
    # -------------------------------------------------
    try:
        import matplotlib.pyplot as plt

        # Use recovery_df already computed (exclude the TOTAL row)
        recovery_plot_df = recovery_df[recovery_df["Campaign"] != "TOTAL"]

        fig, ax = plt.subplots()

        # Bar plot (histogram-like) of tracer recovery % per campaign
        ax.bar(
            recovery_plot_df["Campaign"],
            recovery_plot_df["T_rate_%"],
        )

        ax.set_xlabel("Sampling campaign")
        ax.set_ylabel("Tracer recovery (%)")
        ax.set_title("Tracer mass recovery per campaign")
        fig.tight_layout()

        # Save as a new figure (rename if you prefer to overwrite the old one)
        fig.savefig("Tracer_Recovery_Histogram.png", dpi=300)
        plt.close(fig)
        print("📊 Saved plot: Tracer_Recovery_Histogram.png")

    except Exception as e:
        print(f"⚠ WARNING: Could not create recovery histogram: {e}")

    # -------------------------------------------------
    # 10. Save workbook
    # -------------------------------------------------
    wb.save(output_file)

    print("\n✅ Processing completed successfully.")
    print(f"📁 Output saved as: {output_file}")


if __name__ == "__main__":
    main()


